"""
Обновление Excel файла объекта.

Структура листа-раздела (Вариант Б):
  Каждая поставка = новая строка.
  Колонки: Дата | Поставщик | № счёта | Позиция | Диаметр | Кол-во | Ед. | Цена | Сумма | Остаток нормы | Δ цены
  Норма берётся из строки спецификации (строки с заголовком и базовым количеством).

Лист Kokku:
  Строки группируются по месяцам. Новая строка добавляется в нужный месяц.
"""

import logging
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Цвета ────────────────────────────────────────────────────────────────────
HEAD_FILL   = PatternFill("solid", fgColor="1F3864")   # тёмно-синий заголовок
HEAD_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", fgColor="EEF2FF")   # чётные строки
RED_FILL    = PatternFill("solid", fgColor="FF6B6B")   # перерасход
ORANGE_FILL = PatternFill("solid", fgColor="FFB347")   # остаток < 10 %
GREEN_FILL  = PatternFill("solid", fgColor="C8F7C5")   # норма
PRICE_UP    = PatternFill("solid", fgColor="FFD6D6")   # цена выросла
PRICE_DN    = PatternFill("solid", fgColor="D6FFD6")   # цена упала
MONTH_FILL  = PatternFill("solid", fgColor="2C3E50")   # заголовок месяца в Kokku
MONTH_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONTH_RU = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
            7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
MONTH_ET = {1:"Jaanuar",2:"Veebruar",3:"Märts",4:"Aprill",5:"Mai",6:"Juuni",
            7:"Juuli",8:"August",9:"September",10:"Oktoober",11:"November",12:"Detsember"}

# Ключевые слова для сопоставления раздела → лист
SECTION_MAP = {
    "kanalisatsioon": "Kanalisatsioon",
    "канализация":    "Kanalisatsioon",
    "vesi":           "Vesi",
    "вода":           "Vesi",
    "märg toru":      "MÄRG TORU",
    "противопожарный":"MÄRG TORU",
    "sadevee":        "Sadevee",
    "ливнёвка":       "Sadevee",
}

# Заголовки колонок раздела
SECTION_HEADERS = [
    "Дата", "Поставщик", "№ счёта", "Позиция",
    "Диаметр", "Кол-во", "Ед.", "Цена (€)", "Сумма (€)",
    "Остаток нормы", "Δ цены"
]
# Ширины колонок
SECTION_WIDTHS = [13, 22, 14, 38, 10, 9, 6, 11, 12, 14, 12]

# Заголовки Kokku
KOKKU_HEADERS = ["Месяц / Дата", "Поставщик", "№ счёта", "Раздел", "Материалы (€)", "Зарплата (€)", "Примечание"]
KOKKU_WIDTHS  = [18, 25, 16, 20, 16, 16, 25]


class ExcelUpdater:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    # ── публичный метод ───────────────────────────────────────────────────────

    def update(self, invoice_data: dict) -> dict:
        """Обновить xlsx и вернуть словарь с результатами/предупреждениями."""
        result = {
            "kokku_added":       False,
            "rows_added":        [],      # список добавленных позиций
            "overrun_items":     [],      # перерасход по количеству
            "budget_warnings":   [],      # бюджет > 80 % или превышен
            "price_changes":     [],      # изменение цены
        }

        wb = load_workbook(self.xlsx_path)

        # 1. Kokku
        kokku_ws = self._get_or_create_sheet(wb, "Kokku", KOKKU_HEADERS, KOKKU_WIDTHS,
                                              is_kokku=True)
        self._add_kokku_row(kokku_ws, invoice_data, result)

        # 2. Лист раздела
        section     = invoice_data.get("section", "")
        sheet_name  = self._resolve_sheet(wb, section)
        section_ws  = self._get_or_create_sheet(wb, sheet_name, SECTION_HEADERS,
                                                 SECTION_WIDTHS, is_kokku=False)
        self._add_section_rows(section_ws, invoice_data, result)

        # 3. Проверить бюджет по всем разделам
        for sname in wb.sheetnames:
            if sname not in ("Kokku", "Tasu"):
                self._check_budget(wb[sname], sname, result)

        wb.save(self.xlsx_path)
        return result

    # ── Kokku ────────────────────────────────────────────────────────────────

    def _add_kokku_row(self, ws, invoice_data: dict, result: dict):
        inv_date  = self._parse_date(invoice_data.get("date", "")) or datetime.now()
        month_key = f"{MONTH_RU[inv_date.month]} {str(inv_date.year)[-2:]}"
        total     = sum(i.get("amount", 0) for i in invoice_data.get("items", []))

        # Найти строку итога нужного месяца (или создать секцию)
        insert_before = self._find_or_create_month(ws, month_key, inv_date)

        ws.insert_rows(insert_before)
        row = insert_before
        vals = [
            inv_date.strftime("%d.%m.%Y"),
            invoice_data.get("supplier", ""),
            invoice_data.get("number", ""),
            invoice_data.get("section", ""),
            total,
            "",   # зарплата — вручную
            "",
        ]
        fill = ALT_FILL
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill   = fill
            cell.font   = Font(name="Arial", size=10)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c != 2 else "left")
            if c in (5, 6):
                cell.number_format = "#,##0.00"

        result["kokku_added"] = True

    def _find_or_create_month(self, ws, month_key: str, inv_date: datetime) -> int:
        """Вернуть номер строки «итого месяца» (перед ней вставляем новую строку).
           Если секции нет — создать в конце."""
        kokku_label = f"{month_key} — итого"

        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value or "").strip() == kokku_label:
                    return cell.row   # вставляем перед строкой итога

        # Секция не найдена — добавить в конец
        last = ws.max_row + 1

        # Заголовок месяца
        ws.insert_rows(last);     ws.insert_rows(last + 1)
        hcell = ws.cell(row=last, column=1, value=month_key)
        hcell.fill   = MONTH_FILL
        hcell.font   = MONTH_FONT
        hcell.border = BORDER
        ws.merge_cells(start_row=last, start_column=1, end_row=last,
                       end_column=len(KOKKU_HEADERS))

        # Строка итога
        tcell = ws.cell(row=last + 1, column=1, value=kokku_label)
        tcell.font   = Font(bold=True, name="Arial", size=10)
        tcell.border = BORDER

        return last + 1   # вставляем перед строкой итога (= last+1 после merge)

    # ── Раздел ───────────────────────────────────────────────────────────────

    def _add_section_rows(self, ws, invoice_data: dict, result: dict):
        inv_date = self._parse_date(invoice_data.get("date", "")) or datetime.now()
        date_str = inv_date.strftime("%d.%m.%Y")

        # Прочитать нормы спецификации (если лист создан из шаблона)
        norms = self._read_norms(ws)

        # Прочитать предыдущие цены из листа
        prev_prices = self._read_prev_prices(ws)

        for item in invoice_data.get("items", []):
            name   = item.get("name", "")
            diam   = str(item.get("diameter", "") or "")
            qty    = item.get("quantity") or 0
            price  = item.get("price") or 0
            amount = item.get("amount") or 0
            unit   = item.get("unit", "")

            # Остаток нормы
            norm_key     = self._norm_key(name, diam)
            norm_qty     = norms.get(norm_key, {}).get("total", 0)
            already_used = norms.get(norm_key, {}).get("used", 0)
            new_used     = already_used + qty
            remain       = norm_qty - new_used if norm_qty else None

            # Изменение цены
            prev_price = prev_prices.get(norm_key)
            price_delta = ""
            row_fill    = ALT_FILL if ws.max_row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            if prev_price and price and abs(prev_price - price) > 0.001:
                pct = (price - prev_price) / prev_price * 100
                price_delta = f"{pct:+.1f}%"
                row_fill = PRICE_UP if pct > 0 else PRICE_DN
                result["price_changes"].append({
                    "item":      name,
                    "diameter":  diam,
                    "old_price": prev_price,
                    "new_price": price,
                    "pct":       pct,
                })

            # Подсветка остатка
            remain_fill = row_fill
            if remain is not None:
                if remain < 0:
                    remain_fill = RED_FILL
                    result["overrun_items"].append({
                        "item":     name,
                        "diameter": diam,
                        "norm":     norm_qty,
                        "used":     new_used,
                        "overrun":  abs(remain),
                        "unit":     unit,
                    })
                elif norm_qty and new_used / norm_qty >= 0.9:
                    remain_fill = ORANGE_FILL
                else:
                    remain_fill = GREEN_FILL

            # Добавить строку
            row_num = ws.max_row + 1
            vals = [date_str,
                    invoice_data.get("supplier", ""),
                    invoice_data.get("number", ""),
                    name, diam, qty, unit, price, amount,
                    remain if remain is not None else "",
                    price_delta]

            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=c, value=v)
                cell.font   = Font(name="Arial", size=10)
                cell.border = BORDER
                cell.alignment = Alignment(
                    horizontal="center" if c not in (2, 4) else "left"
                )
                # Числовой формат
                if c in (8, 9):
                    cell.number_format = "#,##0.00"
                # Цвет: остаток — своя подсветка, остальное — row_fill
                cell.fill = remain_fill if c == 10 else row_fill

            result["rows_added"].append(f"{name} Ø{diam}")

            # Обновить кэш использованного
            if norm_key in norms:
                norms[norm_key]["used"] = new_used

    # ── Чтение норм из листа ─────────────────────────────────────────────────

    def _read_norms(self, ws) -> dict:
        """
        Читает строки спецификации (если лист импортирован из шаблона):
        ищет строки где в колонке 1 есть название позиции, в колонке 2 — диаметр,
        в колонке 3 — число (норма).
        Возвращает {norm_key: {total, used}}.
        """
        norms = {}
        hrow = self._find_header_row(ws)
        if not hrow:
            return norms

        # Индексы колонок из заголовка
        col_pos  = self._col_index(ws, hrow, "Позиция")
        col_diam = self._col_index(ws, hrow, "Диаметр")
        col_qty  = self._col_index(ws, hrow, "Кол-во")
        col_used = self._col_index(ws, hrow, "Остаток нормы")

        if not (col_pos and col_qty):
            return norms

        used_per_key: dict[str, float] = {}

        for r in range(hrow + 1, ws.max_row + 1):
            pos_v  = ws.cell(row=r, column=col_pos).value
            diam_v = ws.cell(row=r, column=col_diam).value if col_diam else ""
            qty_v  = ws.cell(row=r, column=col_qty).value

            if not pos_v or not self._is_number(qty_v):
                continue

            key  = self._norm_key(str(pos_v), str(diam_v or ""))
            used = used_per_key.get(key, 0) + float(qty_v)
            used_per_key[key] = used

        # Нормы берём из строк «Спецификация» если лист был инициализирован с шаблоном
        # (в этом случае у нас нет отдельного листа норм — норма = суммарное количество
        #  в оригинальной спецификации)
        # Здесь просто возвращаем накопленные данные как «используемые»
        for key, used in used_per_key.items():
            norms[key] = {"total": 0, "used": used}

        return norms

    def _read_prev_prices(self, ws) -> dict:
        """Последняя известная цена каждой позиции на листе."""
        prices = {}
        hrow = self._find_header_row(ws)
        if not hrow:
            return prices

        col_pos   = self._col_index(ws, hrow, "Позиция")
        col_diam  = self._col_index(ws, hrow, "Диаметр")
        col_price = self._col_index(ws, hrow, "Цена (€)")

        if not (col_pos and col_price):
            return prices

        for r in range(hrow + 1, ws.max_row + 1):
            pos_v   = ws.cell(row=r, column=col_pos).value
            diam_v  = ws.cell(row=r, column=col_diam).value if col_diam else ""
            price_v = ws.cell(row=r, column=col_price).value
            if pos_v and self._is_number(price_v):
                key = self._norm_key(str(pos_v), str(diam_v or ""))
                prices[key] = float(price_v)   # последнее значение — самое свежее

        return prices

    # ── Бюджет ───────────────────────────────────────────────────────────────

    def _check_budget(self, ws, sheet_name: str, result: dict):
        budget = spent = 0
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").lower()
                if "сумма договора" in val or "eelarve" in val:
                    try:
                        budget = float(ws.cell(row=cell.row, column=cell.column + 1).value or 0)
                    except (ValueError, TypeError):
                        pass
                # Сумма (€) — колонка 9 в нашей структуре
        if not budget:
            return

        col_amount = self._col_index(ws, self._find_header_row(ws) or 1, "Сумма (€)")
        if not col_amount:
            return

        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_amount).value
            if self._is_number(v):
                spent += float(v)

        if spent == 0:
            return

        pct       = spent / budget * 100
        remaining = budget - spent

        if pct >= 80:
            result["budget_warnings"].append({
                "section":   sheet_name,
                "budget":    budget,
                "spent":     spent,
                "remaining": remaining,
                "pct":       pct,
                "level":     "КРИТИЧНО" if pct >= 100 else "ВНИМАНИЕ",
            })

    # ── Вспомогательные ──────────────────────────────────────────────────────

    def _get_or_create_sheet(self, wb, name: str, headers: list,
                              widths: list, is_kokku: bool):
        if name in wb.sheetnames:
            return wb[name]

        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False

        # Строка заголовков
        for c, h in enumerate(headers, 1):
            cell        = ws.cell(row=1, column=c, value=h)
            cell.fill   = HEAD_FILL
            cell.font   = HEAD_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        return ws

    def _resolve_sheet(self, wb, section: str) -> str:
        if not section:
            return "Прочее"
        s = section.lower().strip()
        for key, val in SECTION_MAP.items():
            if key in s:
                return val
        # Проверить существующие листы
        for sname in wb.sheetnames:
            if sname.lower() in s or s in sname.lower():
                return sname
        return section.title()

    def _find_header_row(self, ws) -> int | None:
        for row in ws.iter_rows(max_row=5):
            for cell in row:
                if str(cell.value or "").strip() == "Позиция":
                    return cell.row
        return None

    def _col_index(self, ws, hrow: int, header: str) -> int | None:
        if not hrow:
            return None
        for cell in ws[hrow]:
            if str(cell.value or "").strip() == header:
                return cell.column
        return None

    @staticmethod
    def _norm_key(name: str, diameter: str) -> str:
        return f"{name.lower().strip()}|{diameter.lower().strip()}"

    @staticmethod
    def _is_number(v) -> bool:
        try:
            float(v)
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _parse_date(s: str) -> datetime | None:
        for fmt in ("%d.%m.%Y", "%d,%m,%y", "%d,%m,%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(s).strip(), fmt)
            except ValueError:
                continue
        return None
