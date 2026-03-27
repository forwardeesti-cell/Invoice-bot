"""
Экспорт аналитики в Excel
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", name='Arial', size=10)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")
ORANGE_FILL = PatternFill("solid", fgColor="FFE5CC")
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


class ExcelExporter:
    def __init__(self, db):
        self.db = db

    def export(self, object_name: str = None) -> str:
        """Создать Excel с аналитикой. object_name=None — все объекты"""
        wb = Workbook()
        wb.remove(wb.active)

        if object_name:
            objects = [self.db.get_object_by_name(object_name)]
            objects = [o for o in objects if o]
        else:
            objects = self.db.get_all_objects()

        if not objects:
            return None

        # Сводный лист
        self._create_summary_sheet(wb, objects)

        # Лист по каждому объекту
        for obj in objects:
            self._create_object_sheet(wb, obj['name'])

        # Лист истории цен
        self._create_price_history_sheet(wb, object_name)

        out_path = f"/tmp/analytics_{object_name or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(out_path)
        return out_path

    def _create_summary_sheet(self, wb, objects: list):
        ws = wb.create_sheet("Сводка")
        ws.sheet_view.showGridLines = False

        title_font = Font(bold=True, size=14, name='Arial', color='2C3E50')
        ws['A1'] = "📊 Аналитика по объектам"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='left')
        ws.merge_cells('A1:F1')
        ws['A2'] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = Font(italic=True, size=9, color='666666', name='Arial')

        headers = ['Объект', 'Кол-во счетов', 'Сумма материалов (€)', 'Файл Excel', 'Статус']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER

        for i, obj in enumerate(objects):
            row = 5 + i
            fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            data = [obj['name'], obj.get('invoice_count', 0),
                    obj.get('total', 0), obj.get('xlsx_path', '—'), '✅ Активен']
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
                cell.border = BORDER
                if col == 3:
                    cell.number_format = '#,##0.00'

        col_widths = [30, 18, 25, 40, 15]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _create_object_sheet(self, wb, object_name: str):
        safe_name = object_name[:28].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(safe_name)
        ws.sheet_view.showGridLines = False

        ws['A1'] = f"Объект: {object_name}"
        ws['A1'].font = Font(bold=True, size=13, name='Arial', color='2C3E50')
        ws.merge_cells('A1:G1')

        headers = ['Дата', 'Поставщик', 'Номер счёта', 'Раздел', 'Позиция', 'Кол-во', 'Сумма (€)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER

        # Получить данные из БД
        obj = self.db.get_object_by_name(object_name)
        if not obj:
            return

        cur = self.db.conn.execute("""
            SELECT i.date, i.supplier, i.number, ii.section, ii.name, ii.quantity, ii.amount
            FROM invoices i JOIN invoice_items ii ON ii.invoice_id = i.id
            WHERE i.object_id = ? ORDER BY i.date, i.supplier
        """, (obj['id'],))
        rows = cur.fetchall()

        for i, row in enumerate(rows):
            r = 4 + i
            fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center' if col != 5 else 'left')
                cell.border = BORDER
                if col == 7:
                    cell.number_format = '#,##0.00'

        # Итого
        if rows:
            total_row = 4 + len(rows)
            ws.cell(row=total_row, column=6, value='ИТОГО:').font = Font(bold=True, name='Arial')
            ws.cell(row=total_row, column=7,
                    value=f'=SUM(G4:G{total_row-1})').number_format = '#,##0.00'
            ws.cell(row=total_row, column=7).font = Font(bold=True, name='Arial')

        col_widths = [14, 25, 18, 20, 40, 10, 16]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _create_price_history_sheet(self, wb, object_name: str = None):
        ws = wb.create_sheet("История цен")
        ws.sheet_view.showGridLines = False

        ws['A1'] = "📈 История изменения цен"
        ws['A1'].font = Font(bold=True, size=13, name='Arial', color='2C3E50')
        ws.merge_cells('A1:G1')

        headers = ['Позиция', 'Диаметр', 'Цена', 'Поставщик', 'Объект', 'Дата счёта', 'Изменение']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER

        query = """
            SELECT item_name, diameter, price, supplier, object_name, invoice_date
            FROM price_history ORDER BY item_name, created_at
        """
        if object_name:
            query = query.replace("ORDER BY", "WHERE object_name=? ORDER BY")
            cur = self.db.conn.execute(query, (object_name,))
        else:
            cur = self.db.conn.execute(query)

        rows = cur.fetchall()
        prev_prices = {}

        for i, row in enumerate(rows):
            r = 4 + i
            item_key = (row['item_name'], row['diameter'])
            prev_price = prev_prices.get(item_key)
            prev_prices[item_key] = row['price']

            change_str = ''
            fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            if prev_price and abs(prev_price - row['price']) > 0.01:
                pct = ((row['price'] - prev_price) / prev_price) * 100
                change_str = f"{pct:+.1f}%"
                fill = RED_FILL if pct > 0 else PatternFill("solid", fgColor="CCFFCC")

            data = [row['item_name'], row['diameter'], row['price'],
                    row['supplier'], row['object_name'], row['invoice_date'], change_str]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center' if col != 1 else 'left')
                cell.border = BORDER
                if col == 3:
                    cell.number_format = '#,##0.00'

        col_widths = [40, 14, 14, 25, 25, 14, 14]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
