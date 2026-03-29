#!/usr/bin/env python3
"""
Invoice Bot — всё в одном файле
"""

import os, json, logging, shutil, sqlite3, base64, re
from pathlib import Path
from datetime import datetime
from typing import Optional

# ── Проверка зависимостей ────────────────────────────────────────────────────
try:
    from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
    from telegram.ext import (Application, CommandHandler, MessageHandler,
                               CallbackQueryHandler, ConversationHandler,
                               filters, ContextTypes)
except ImportError:
    raise SystemExit("Установите: pip install python-telegram-bot==21.6")

try:
    import anthropic
except ImportError:
    raise SystemExit("Установите: pip install anthropic")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Установите: pip install openpyxl")

# ── Логирование ──────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
    handlers=[logging.StreamHandler()]
)
log = logging.getLogger(__name__)

# ── Константы ────────────────────────────────────────────────────────────────
WAITING_OBJECT  = 1
WAITING_SECTION = 2
WAITING_XLSX    = 3

OBJECTS_DIR = Path("objects")
OBJECTS_DIR.mkdir(exist_ok=True)

MONTH_RU = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
            7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
MONTH_ET = {1:"Jaanuar",2:"Veebruar",3:"Märts",4:"Aprill",5:"Mai",6:"Juuni",
            7:"Juuli",8:"August",9:"September",10:"Oktoober",11:"November",12:"Detsember"}

SECTION_MAP = {
    "kanalisatsioon":"Kanalisatsioon","канализация":"Kanalisatsioon",
    "vesi":"Vesi","вода":"Vesi",
    "märg toru":"MÄRG TORU","противопожарный":"MÄRG TORU",
    "sadevee":"Sadevee","ливнёвка":"Sadevee",
}

SECTION_HEADERS = ["Дата","Поставщик","№ счёта","Позиция","Диаметр",
                   "Кол-во","Ед.","Цена (€)","Сумма (€)","Остаток нормы","Δ цены"]
SECTION_WIDTHS  = [13,22,14,38,10,9,6,11,12,14,12]
KOKKU_HEADERS   = ["Месяц / Дата","Поставщик","№ счёта","Раздел","Материалы (€)","Зарплата (€)","Примечание"]
KOKKU_WIDTHS    = [18,25,16,20,16,16,25]

HEAD_FILL   = PatternFill("solid", fgColor="1F3864")
HEAD_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", fgColor="EEF2FF")
RED_FILL    = PatternFill("solid", fgColor="FF6B6B")
ORANGE_FILL = PatternFill("solid", fgColor="FFB347")
GREEN_FILL  = PatternFill("solid", fgColor="C8F7C5")
PRICE_UP    = PatternFill("solid", fgColor="FFD6D6")
PRICE_DN    = PatternFill("solid", fgColor="D6FFD6")
MONTH_FILL  = PatternFill("solid", fgColor="2C3E50")
MONTH_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ════════════════════════════════════════════════════════════════════════════
#  БАЗА ДАННЫХ
# ════════════════════════════════════════════════════════════════════════════

class DB:
    def __init__(self):
        self.con = sqlite3.connect("invoices.db", check_same_thread=False)
        self.con.row_factory = sqlite3.Row
        self.con.executescript("""
        CREATE TABLE IF NOT EXISTS objects(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            xlsx_path TEXT,
            created_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS invoices(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            object_id INTEGER, number TEXT, date TEXT,
            supplier TEXT, total REAL, section TEXT, raw_json TEXT,
            created_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS invoice_items(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_id INTEGER, name TEXT, diameter TEXT,
            quantity REAL, unit TEXT, price REAL, amount REAL,
            category TEXT, section TEXT);
        CREATE TABLE IF NOT EXISTS price_history(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_name TEXT, diameter TEXT, price REAL,
            supplier TEXT, object_name TEXT, invoice_date TEXT,
            created_at TEXT DEFAULT(datetime('now')));
        """)
        self.con.commit()

    def q(self, sql, params=()):
        return self.con.execute(sql, params)

    def get_or_create_object(self, name, xlsx=None):
        r = self.q("SELECT id FROM objects WHERE name=?", (name,)).fetchone()
        if r:
            if xlsx:
                self.q("UPDATE objects SET xlsx_path=? WHERE id=?", (xlsx, r["id"]))
                self.con.commit()
            return r["id"]
        cur = self.q("INSERT INTO objects(name,xlsx_path) VALUES(?,?)", (name, xlsx))
        self.con.commit()
        return cur.lastrowid

    def save_invoice(self, inv):
        oid = self.get_or_create_object(inv.get("object","?"))
        total = sum(i.get("amount",0) for i in inv.get("items",[]))
        cur = self.q(
            "INSERT INTO invoices(object_id,number,date,supplier,total,section,raw_json) VALUES(?,?,?,?,?,?,?)",
            (oid, inv.get("number"), inv.get("date"), inv.get("supplier"),
             total, inv.get("section"), json.dumps(inv, ensure_ascii=False)))
        iid = cur.lastrowid
        for item in inv.get("items",[]):
            self.q("INSERT INTO invoice_items(invoice_id,name,diameter,quantity,unit,price,amount,category,section) VALUES(?,?,?,?,?,?,?,?,?)",
                   (iid, item.get("name"), item.get("diameter"), item.get("quantity"),
                    item.get("unit"), item.get("price"), item.get("amount"),
                    item.get("category"), item.get("section")))
            if item.get("price"):
                self.q("INSERT INTO price_history(item_name,diameter,price,supplier,object_name,invoice_date) VALUES(?,?,?,?,?,?)",
                       (item["name"], item.get("diameter"), item["price"],
                        inv.get("supplier"), inv.get("object"), inv.get("date")))
        self.con.commit()

    def get_objects(self):
        return [dict(r) for r in self.q("""
            SELECT o.name, o.xlsx_path,
                   COUNT(DISTINCT i.id) as invoice_count,
                   COALESCE(SUM(i.total),0) as total
            FROM objects o LEFT JOIN invoices i ON i.object_id=o.id
            GROUP BY o.id ORDER BY o.name""").fetchall()]

    def get_xlsx(self, name):
        r = self.q("SELECT xlsx_path FROM objects WHERE name=?", (name,)).fetchone()
        return r["xlsx_path"] if r else None

    def set_xlsx(self, name, path):
        self.get_or_create_object(name, path)
        self.q("UPDATE objects SET xlsx_path=? WHERE name=?", (path, name))
        self.con.commit()

    def get_report(self, name):
        r = self.q("SELECT id FROM objects WHERE name=?", (name,)).fetchone()
        if not r: return None
        oid = r["id"]
        s = self.q("SELECT COUNT(*) as c, COALESCE(SUM(total),0) as t FROM invoices WHERE object_id=?", (oid,)).fetchone()
        cats = {r["section"]: r["a"] for r in self.q(
            "SELECT section, COALESCE(SUM(amount),0) as a FROM invoice_items ii JOIN invoices i ON ii.invoice_id=i.id WHERE i.object_id=? GROUP BY section ORDER BY a DESC", (oid,)).fetchall()}
        recent = [dict(r) for r in self.q(
            "SELECT number,date,total FROM invoices WHERE object_id=? ORDER BY created_at DESC LIMIT 5", (oid,)).fetchall()]
        return {"count": s["c"], "total": s["t"], "cats": cats, "recent": recent}

db = DB()

# ════════════════════════════════════════════════════════════════════════════
#  ИЗВЛЕЧЕНИЕ PDF
# ════════════════════════════════════════════════════════════════════════════

async def extract_pdf(pdf_path: str) -> Optional[dict]:
    api_key = os.environ.get("ANTHROPIC_API_KEY","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")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY не задан!")
    client = anthropic.Anthropic(api_key=api_key)
    with open(pdf_path,"rb") as f:
        pdf_b64 = base64.standard_b64encode(f.read()).decode()
    prompt = """Проанализируй PDF счёт и верни ТОЛЬКО JSON без markdown:
{
  "number": "номер счёта или null",
  "date": "дата DD.MM.YYYY или null",
  "object": "объект/адрес или null",
  "supplier": "поставщик или null",
  "section": "раздел: Kanalisatsioon/Vesi/MÄRG TORU/Sadevee или null",
  "items": [
    {
      "name": "название позиции",
      "diameter": "диаметр или null",
      "quantity": число,
      "unit": "ед.изм.",
      "price": цена за единицу,
      "amount": итоговая сумма позиции
    }
  ],
  "total": общая сумма
}
Верни ТОЛЬКО JSON, никакого другого текста."""
    try:
        resp = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=4000,
            messages=[{"role":"user","content":[
                {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":pdf_b64}},
                {"type":"text","text":prompt}
            ]}]
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*","",raw)
        raw = re.sub(r"\s*```$","",raw)
        return json.loads(raw)
    except Exception as e:
        log.error(f"PDF extract error: {e}")
        return None

# ════════════════════════════════════════════════════════════════════════════
#  ОБНОВЛЕНИЕ EXCEL
# ════════════════════════════════════════════════════════════════════════════

def parse_date(s):
    for fmt in ("%d.%m.%Y","%d,%m,%y","%d,%m,%Y","%Y-%m-%d","%d/%m/%Y"):
        try: return datetime.strptime(str(s or "").strip(), fmt)
        except ValueError: pass
    return None

def is_num(v):
    try: float(v); return True
    except: return False

def make_sheet(wb, name, headers, widths):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(row=1,column=c,value=h)
        cell.fill=HEAD_FILL; cell.font=HEAD_FONT; cell.border=BORDER
        cell.alignment=Alignment(horizontal="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=28
    ws.freeze_panes="A2"
    return ws

def resolve_sheet(wb, section):
    if not section: return "Прочее"
    s = section.lower().strip()
    for k,v in SECTION_MAP.items():
        if k in s: return v
    for sn in wb.sheetnames:
        if sn.lower() in s or s in sn.lower(): return sn
    return section.title()

def find_or_create_month(ws, month_key):
    kokku_label = f"{month_key} — итого"
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip() == kokku_label:
                return cell.row
    last = ws.max_row + 1
    ws.insert_rows(last); ws.insert_rows(last+1)
    h = ws.cell(row=last,column=1,value=month_key)
    h.fill=MONTH_FILL; h.font=MONTH_FONT; h.border=BORDER
    ws.merge_cells(start_row=last,start_column=1,end_row=last,end_column=len(KOKKU_HEADERS))
    t = ws.cell(row=last+1,column=1,value=kokku_label)
    t.font=Font(bold=True,name="Arial",size=10); t.border=BORDER
    return last+1

def update_excel(xlsx_path: str, inv: dict) -> dict:
    result = {"kokku_added":False,"rows_added":0,"overrun_items":[],
              "budget_warnings":[],"price_changes":[]}
    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        return {"error": str(e)}

    inv_date = parse_date(inv.get("date","")) or datetime.now()
    month_key = f"{MONTH_RU[inv_date.month]} {str(inv_date.year)[-2:]}"
    total = sum(i.get("amount",0) for i in inv.get("items",[]))

    # ── Kokku ────────────────────────────────────────────────────────────────
    if "Kokku" not in wb.sheetnames:
        make_sheet(wb,"Kokku",KOKKU_HEADERS,KOKKU_WIDTHS)
    ws_k = wb["Kokku"]
    ins = find_or_create_month(ws_k, month_key)
    ws_k.insert_rows(ins)
    for c,v in enumerate([inv_date.strftime("%d.%m.%Y"), inv.get("supplier",""),
                           inv.get("number",""), inv.get("section",""), total,"",""],1):
        cell=ws_k.cell(row=ins,column=c,value=v)
        cell.fill=ALT_FILL; cell.font=Font(name="Arial",size=10); cell.border=BORDER
        cell.alignment=Alignment(horizontal="center" if c!=2 else "left")
        if c==5: cell.number_format="#,##0.00"
    result["kokku_added"]=True

    # ── Раздел ───────────────────────────────────────────────────────────────
    sheet_name = resolve_sheet(wb, inv.get("section",""))
    if sheet_name not in wb.sheetnames:
        make_sheet(wb,sheet_name,SECTION_HEADERS,SECTION_WIDTHS)
    ws_s = wb[sheet_name]

    # Предыдущие цены
    prev_prices = {}
    hrow = None
    for row in ws_s.iter_rows(max_row=5):
        for cell in row:
            if str(cell.value or "").strip()=="Позиция":
                hrow=cell.row; break
        if hrow: break

    if hrow:
        col_p=col_d=col_pr=None
        for cell in ws_s[hrow]:
            v=str(cell.value or "").strip()
            if v=="Позиция": col_p=cell.column
            elif v=="Диаметр": col_d=cell.column
            elif v=="Цена (€)": col_pr=cell.column
        if col_p and col_pr:
            for r in range(hrow+1, ws_s.max_row+1):
                pv=ws_s.cell(row=r,column=col_p).value
                dv=ws_s.cell(row=r,column=col_d).value if col_d else ""
                prv=ws_s.cell(row=r,column=col_pr).value
                if pv and is_num(prv):
                    key=f"{str(pv).lower()}|{str(dv or '').lower()}"
                    prev_prices[key]=float(prv)

    for item in inv.get("items",[]):
        name=item.get("name",""); diam=str(item.get("diameter","") or "")
        qty=item.get("quantity") or 0; price=item.get("price") or 0
        amount=item.get("amount") or 0; unit=item.get("unit","")
        key=f"{name.lower()}|{diam.lower()}"

        prev=prev_prices.get(key)
        delta=""; rfill=ALT_FILL
        if prev and price and abs(prev-price)>0.001:
            pct=(price-prev)/prev*100
            delta=f"{pct:+.1f}%"
            rfill=PRICE_UP if pct>0 else PRICE_DN
            result["price_changes"].append({"item":name,"diameter":diam,
                "old":prev,"new":price,"pct":pct})

        row_n=ws_s.max_row+1
        for c,v in enumerate([inv_date.strftime("%d.%m.%Y"),
                               inv.get("supplier",""), inv.get("number",""),
                               name, diam, qty, unit, price, amount, "", delta],1):
            cell=ws_s.cell(row=row_n,column=c,value=v)
            cell.font=Font(name="Arial",size=10); cell.border=BORDER
            cell.alignment=Alignment(horizontal="center" if c not in(2,4) else "left")
            if c in(8,9): cell.number_format="#,##0.00"
            cell.fill=rfill
        result["rows_added"]+=1

    # ── Бюджет ───────────────────────────────────────────────────────────────
    for sn in wb.sheetnames:
        if sn in("Kokku","Tasu"): continue
        ws_b=wb[sn]
        budget=0
        for row in ws_b.iter_rows():
            for cell in row:
                val=str(cell.value or "").lower()
                if "сумма договора" in val or "eelarve" in val:
                    try: budget=float(ws_b.cell(row=cell.row,column=cell.column+1).value or 0)
                    except: pass
        if not budget: continue
        spent=0
        for r in range(2,ws_b.max_row+1):
            # Колонка 9 = Сумма (€)
            v=ws_b.cell(row=r,column=9).value
            if is_num(v): spent+=float(v)
        if not spent: continue
        pct=spent/budget*100
        if pct>=80:
            result["budget_warnings"].append({
                "section":sn,"budget":budget,"spent":spent,
                "remaining":budget-spent,"pct":pct,
                "level":"КРИТИЧНО" if pct>=100 else "ВНИМАНИЕ"})

    wb.save(xlsx_path)
    return result

# ════════════════════════════════════════════════════════════════════════════
#  TELEGRAM БОТ
# ════════════════════════════════════════════════════════════════════════════

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Привет! Я бот для учёта счетов.*\n\n"
        "🗂 *Как начать:*\n"
        "1. /setxlsx — загрузите Excel файл объекта\n"
        "2. Отправьте PDF счёт — бот обновит Excel\n\n"
        "📊 *Команды:*\n"
        "/objects — список объектов\n"
        "/report — отчёт по объекту\n"
        "/setxlsx — привязать Excel\n"
        "/help — справка", parse_mode="Markdown")

async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📖 *Справка*\n\n"
        "• Отправьте PDF → бот извлечёт данные и обновит Excel\n"
        "• /setxlsx → отправьте .xlsx → выберите объект\n\n"
        "*Предупреждения:*\n"
        "🔴 Перерасход по материалу\n"
        "📈 Цена выросла\n"
        "⚠️ Бюджет > 80%\n"
        "🚨 Бюджет превышен", parse_mode="Markdown")

async def cmd_objects(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    objs = db.get_objects()
    if not objs:
        await update.message.reply_text("📭 Объектов нет. Загрузите Excel через /setxlsx")
        return
    text = "🏗 *Объекты:*\n\n"
    for o in objs:
        xl = "✅" if o.get("xlsx_path") else "❌"
        text += f"• *{o['name']}* {xl} | Счетов: {o['invoice_count']} | {o['total']:,.2f} €\n"
    await update.message.reply_text(text, parse_mode="Markdown")

async def cmd_report(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    objs = db.get_objects()
    if not objs:
        await update.message.reply_text("📭 Нет объектов.")
        return
    kb = [[InlineKeyboardButton(o["name"], callback_data=f"rep_{o['name']}")] for o in objs]
    await update.message.reply_text("Выберите объект:", reply_markup=InlineKeyboardMarkup(kb))

async def cmd_setxlsx(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["wait_xlsx"] = True
    await update.message.reply_text("📎 Отправьте .xlsx файл объекта:")
    return WAITING_XLSX

async def handle_xlsx(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc.file_name.endswith(".xlsx"):
        await update.message.reply_text("❌ Нужен .xlsx файл")
        return WAITING_XLSX
    file = await doc.get_file()
    tmp = f"/tmp/{doc.file_name}"
    await file.download_to_drive(tmp)
    ctx.user_data["pending_xlsx"] = tmp
    ctx.user_data["pending_xlsx_name"] = doc.file_name
    objs = db.get_objects()
    kb = [[InlineKeyboardButton(o["name"], callback_data=f"xl_{o['name']}")] for o in objs]
    kb.append([InlineKeyboardButton("➕ Новый объект", callback_data="xl_NEW")])
    await update.message.reply_text(
        f"📁 *{doc.file_name}* получен.\nК какому объекту привязать?",
        parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(kb))
    return WAITING_XLSX

async def handle_pdf(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("⏳ Читаю счёт...")
    file = await update.message.document.get_file()
    pdf_path = f"/tmp/inv_{update.message.message_id}.pdf"
    await file.download_to_drive(pdf_path)
    await msg.edit_text("🤖 Анализирую через AI...")
    try:
        inv = await extract_pdf(pdf_path)
    except Exception as e:
        await msg.edit_text(f"❌ Ошибка AI: {e}")
        return ConversationHandler.END
    if not inv:
        await msg.edit_text("❌ Не удалось извлечь данные из файла.")
        return ConversationHandler.END
    ctx.user_data.update({"invoice": inv, "pdf_path": pdf_path, "msg": msg.message_id})
    if not inv.get("object"):
        objs = db.get_objects()
        kb = [[InlineKeyboardButton(o["name"], callback_data=f"io_{o['name']}")] for o in objs]
        kb.append([InlineKeyboardButton("➕ Новый объект", callback_data="io_NEW")])
        await msg.edit_text(
            f"📄 *№{inv.get('number','?')}* от {inv.get('date','?')}\n"
            f"🏢 {inv.get('supplier','?')}\n\n❓ *К какому объекту?*",
            parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(kb))
        return WAITING_OBJECT
    return await ask_section_or_save(update, ctx, msg)

async def ask_section_or_save(upd, ctx, msg):
    inv = ctx.user_data["invoice"]
    if not inv.get("section"):
        secs = ["Kanalisatsioon","Vesi","MÄRG TORU","Sadevee","Прочее"]
        kb = [[InlineKeyboardButton(s, callback_data=f"is_{s}")] for s in secs]
        txt = (f"📄 *№{inv.get('number','?')}* | {inv.get('object','?')}\n"
               f"💰 {sum(i.get('amount',0) for i in inv.get('items',[])):,.2f} €\n\n"
               f"❓ *Раздел:*")
        fn = getattr(msg,"edit_message_text",None) or getattr(msg,"edit_text",None)
        if fn:
            await fn(txt, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(kb))
        else:
            await upd.message.reply_text(txt, parse_mode="Markdown",
                                          reply_markup=InlineKeyboardMarkup(kb))
        return WAITING_SECTION
    await finalize(upd, ctx, msg)
    return ConversationHandler.END

async def finalize(upd, ctx, msg):
    inv      = ctx.user_data["invoice"]
    obj_name = inv.get("object","?")
    total    = sum(i.get("amount",0) for i in inv.get("items",[]))

    db.save_invoice(inv)

    xlsx_path = db.get_xlsx(obj_name)
    er = {}
    if xlsx_path and Path(xlsx_path).exists():
        er = update_excel(xlsx_path, inv)

    # Основной текст
    text = (f"✅ *Счёт обработан*\n\n"
            f"📄 №*{inv.get('number','?')}* от {inv.get('date','?')}\n"
            f"🏗 Объект: *{obj_name}*\n"
            f"📂 Раздел: {inv.get('section','?')}\n"
            f"🏢 Поставщик: {inv.get('supplier','?')}\n"
            f"📦 Позиций: {len(inv.get('items',[]))}  |  💰 *{total:,.2f} €*\n")

    if not xlsx_path:
        text += "\n💡 _Привяжите Excel через /setxlsx_\n"
    elif er.get("error"):
        text += f"\n⚠️ _Ошибка Excel: {er['error']}_\n"
    else:
        text += f"\n📊 Excel обновлён ✅  |  Строк добавлено: {er.get('rows_added',0)}\n"

    if er.get("overrun_items"):
        text += "\n🔴 *ПЕРЕРАСХОД:*\n"
        for w in er["overrun_items"]:
            text += f"  ❗ *{w['item']}* Ø{w['diameter']}: -{w['overrun']:.1f} {w['unit']}\n"

    if er.get("budget_warnings"):
        text += "\n"
        for w in er["budget_warnings"]:
            e = "🚨" if w["level"]=="КРИТИЧНО" else "⚠️"
            text += (f"{e} *{w['section']}*: {w['pct']:.0f}% бюджета\n"
                     f"   Остаток: *{w['remaining']:,.2f} €*\n")

    if er.get("price_changes"):
        text += "\n📈 *Изменение цен:*\n"
        for ch in er["price_changes"]:
            a = "📈" if ch["pct"]>0 else "📉"
            text += (f"  {a} *{ch['item']}*"
                     f"{' Ø'+ch['diameter'] if ch['diameter'] else ''}: "
                     f"{ch['old']:,.2f}→{ch['new']:,.2f} € (*{ch['pct']:+.1f}%*)\n")

    fn = (getattr(msg,"edit_message_text",None) or getattr(msg,"edit_text",None))
    try:
        if fn: await fn(text, parse_mode="Markdown")
        else:  await upd.message.reply_text(text, parse_mode="Markdown")
    except Exception:
        if hasattr(upd,"message"):
            await upd.message.reply_text(text, parse_mode="Markdown")

    if xlsx_path and Path(xlsx_path).exists() and not er.get("error"):
        chat_id = upd.message.chat_id if hasattr(upd,"message") else upd.effective_chat.id
        with open(xlsx_path,"rb") as f:
            await ctx.bot.send_document(chat_id, f,
                caption=f"📊 {obj_name} — обновлён", filename=Path(xlsx_path).name)

    pdf = ctx.user_data.get("pdf_path")
    ctx.user_data.clear()
    if pdf: Path(pdf).unlink(missing_ok=True)

async def on_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    d = q.data

    # Привязка xlsx
    if d.startswith("xl_"):
        obj = d[3:]
        if obj == "NEW":
            await q.edit_message_text("Введите название нового объекта:")
            ctx.user_data["new_obj_xlsx"] = True
            return WAITING_XLSX
        tmp  = ctx.user_data.get("pending_xlsx")
        fname= ctx.user_data.get("pending_xlsx_name","object.xlsx")
        if tmp and Path(tmp).exists():
            dest = OBJECTS_DIR / f"{obj}_{fname}"
            shutil.copy2(tmp, dest); Path(tmp).unlink(missing_ok=True)
            db.set_xlsx(obj, str(dest))
            await q.edit_message_text(f"✅ Excel привязан к *{obj}*", parse_mode="Markdown")
        else:
            await q.edit_message_text("❌ Файл не найден.")
        ctx.user_data.clear()
        return ConversationHandler.END

    # Объект для счёта
    if d.startswith("io_"):
        obj = d[3:]
        if obj == "NEW":
            await q.edit_message_text("Введите название объекта:")
            ctx.user_data["new_inv_obj"] = True
            return WAITING_OBJECT
        ctx.user_data["invoice"]["object"] = obj
        return await ask_section_or_save(q, ctx, q)

    # Раздел
    if d.startswith("is_"):
        ctx.user_data["invoice"]["section"] = d[3:]
        await finalize(q, ctx, q)
        return ConversationHandler.END

    # Отчёт
    if d.startswith("rep_"):
        name = d[4:]
        r = db.get_report(name)
        if not r:
            await q.edit_message_text("❌ Не найден.")
            return
        text = f"📊 *{name}*\n\nСчетов: {r['count']} | {r['total']:,.2f} €\n"
        if r["cats"]:
            text += "\n📂 *Разделы:*\n"
            for s,a in r["cats"].items():
                text += f"  • {s or '?'}: {a:,.2f} €\n"
        if r["recent"]:
            text += "\n🗓 *Последние:*\n"
            for i in r["recent"]:
                text += f"  • №{i['number']} {i['date']} — {i['total']:,.2f} €\n"
        await q.edit_message_text(text, parse_mode="Markdown")

async def on_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    # Новый объект для xlsx
    if ctx.user_data.get("new_obj_xlsx"):
        ctx.user_data["new_obj_xlsx"] = False
        tmp  = ctx.user_data.get("pending_xlsx")
        fname= ctx.user_data.get("pending_xlsx_name","object.xlsx")
        if tmp and Path(tmp).exists():
            dest = OBJECTS_DIR / f"{text}_{fname}"
            shutil.copy2(tmp, dest); Path(tmp).unlink(missing_ok=True)
            db.set_xlsx(text, str(dest))
            await update.message.reply_text(f"✅ Excel привязан к *{text}*", parse_mode="Markdown")
        ctx.user_data.clear()
        return ConversationHandler.END
    # Новый объект для счёта
    if ctx.user_data.get("new_inv_obj"):
        ctx.user_data["new_inv_obj"] = False
        ctx.user_data["invoice"]["object"] = text
        msg = await update.message.reply_text("✅")
        return await ask_section_or_save(update, ctx, msg)

# ════════════════════════════════════════════════════════════════════════════
#  ЗАПУСК
# ════════════════════════════════════════════════════════════════════════════

def main():
    token = os.environ.get("TELEGRAM_BOT_TOKEN","8693681834:AAF6wBi-y0gRCbLI5R44BJQECnJAZ-HjflM")
    if not token:
        raise SystemExit("❌ Не задан TELEGRAM_BOT_TOKEN!")

    app = Application.builder().token(token).build()

    conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Document.PDF, handle_pdf),
            CommandHandler("setxlsx", cmd_setxlsx),
        ],
        states={
            WAITING_OBJECT:  [CallbackQueryHandler(on_callback, pattern="^io_"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, on_text)],
            WAITING_SECTION: [CallbackQueryHandler(on_callback, pattern="^is_")],
            WAITING_XLSX:    [MessageHandler(filters.Document.ALL, handle_xlsx),
                               CallbackQueryHandler(on_callback, pattern="^xl_"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, on_text)],
        },
        fallbacks=[CommandHandler("start", cmd_start)],
        allow_reentry=True,
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("start",   cmd_start))
    app.add_handler(CommandHandler("help",    cmd_help))
    app.add_handler(CommandHandler("objects", cmd_objects))
    app.add_handler(CommandHandler("report",  cmd_report))
    app.add_handler(CallbackQueryHandler(on_callback))

    log.info("🤖 Бот запущен!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
